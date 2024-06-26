import yaml
import json
import csv
import os
import openpyxl


class PrioritizeControls:
    def __init__(self, config):
        # Load configuration
        config = self.load_config(config)
        self.include_details = config['include_details']

        # Input Files
        attack_top_techniques_file = config['prioritized_techniques']
        attack_nist_mapping_file = config['attack_nist_mappings']
        nist_cis_mapping_file = config['nist_cis_mappings']
        new_cjis_nist_controls_file = config['new_cjis_nist_controls']

        # Initiate lists to be used later
        self.new_cjis_nist_controls = self.load_new_cjis_nist_controls(new_cjis_nist_controls_file)
        self.attack_priorities = self.load_attack_priorities(attack_top_techniques_file)
        self.attack_nist_mapping = self.load_attack_nist_mapping(attack_nist_mapping_file)
        self.nist_cis_mapping = self.load_nist_cis_mapping(nist_cis_mapping_file)

        # Results files
        results_dir = config["results_directory"]
        self.directory_setup(results_dir)
        if self.include_details:
            self.file_ending = "_details.csv"
        else:
            self.file_ending = ".cvs"
        self.attack_priorities_file = os.path.join(results_dir, f"attack_priorities{self.file_ending}")
        self.attack_priorities_with_nist_file = os.path.join(results_dir, f"attack_priorities_with_nist{self.file_ending}")
        self.nist_with_techniques_file = os.path.join(results_dir, f"nist_with_techniques_file{self.file_ending}")

    def load_config(self, config_file):
        """
        Load configurations from provided YAML file.  
        """
        with open(config_file, 'r') as f:
            config = yaml.safe_load(f)    
        return config

    def load_new_cjis_nist_controls(self, new_cjis_nist_controls_file):
        ### Note the following NIST controls are not available in the priority calculator (05/04/2023)
        ### AC-11: Session Lock
        ### AC-14: Permitted Actions Without Identification Or Authentication
        ### IR-5: Incident Monitoring
        ### https://top-attack-techniques.mitre-engenuity.org/calculator
        all_controls = []
        with open(new_cjis_nist_controls_file) as file:
            for line in file:
                remove_whitespace = line.strip()
                if not remove_whitespace: #skip blank lines
                    continue 
                no_leading_zeros = remove_whitespace.replace("-0", "-") # Ensuring formatting matches the input spreadsheets
                all_controls.append(no_leading_zeros)
       
        return all_controls

    def directory_setup(self, results_dir):
        if not os.path.exists(results_dir):
            os.makedirs(results_dir)

    def load_attack_priorities(self, attack_top_techniques_file):
        """
        Load techniques and rankings from output of attack calculator, JSON data
        """
        attack_priorities = []
        with open(attack_top_techniques_file, encoding="utf8") as file:
            json_data = json.load(file)
            for obj in json_data:
                rank = obj["rank"]
                technique = obj["tid"]
                attack_priorities.append((rank, technique))
                subtechniques = obj['subtechniques']
                for sub_obj in subtechniques:
                    # ATT&CK does not assign different priorities to sub techniques
                    subtechnique = sub_obj['tid']
                    attack_priorities.append((rank, subtechnique))

        return attack_priorities

    def load_attack_nist_mapping(self, attack_nist_mapping_file):
        """
        Load NIST 800-53 <-> ATT&CK mapping
        """
        # Load the Excel file
        workbook = openpyxl.load_workbook(attack_nist_mapping_file)

        # Select the worksheet
        worksheet = workbook.active

        # Initialize an empty list to store the CSV data
        attack_nist_mapping = [] 

        # Iterate through each row in the worksheet
        for row in worksheet.iter_rows(min_row=2):
            # Get the values of columns in the current row
            for cell in row:
                if cell.column_letter == 'A':
                    control = cell.value
                # Don't account for sub categories in the mapping, since our other references do no have sub mappings for NIST controls.
                elif cell.column_letter == 'D':
                    technique = cell.value
            
            # Add the values to the data list
            attack_nist_mapping.append((control, technique))
        
        return attack_nist_mapping

    def load_nist_cis_mapping(self, nist_cis_mapping_file):
        # Load the Excel file
        workbook = openpyxl.load_workbook(nist_cis_mapping_file)

        # Select the worksheet
        worksheet = workbook['All CIS Controls & Safeguards']

        # Create a list to store the data
        nist_cis_mapping = []

        # Iterate through each row in the worksheet
        for row in worksheet.iter_rows(min_row=2):
            # Get the values of columns the current row
            for cell in row:
                try:
                    if cell.column_letter == 'B':
                        cis_control = cell.value
                    # Don't account for sub categories in the mapping, since our other references do no have sub mappings for NIST controls.
                    elif cell.column_letter == 'L':
                        nist_mapping= cell.value
                        if nist_mapping and "(" in nist_mapping:
                            nist_mapping = nist_mapping.split("(")[0]
                except AttributeError as e:
                    pass
            
            # Add the values to the data list
            if nist_mapping:
                nist_cis_mapping.append((cis_control, nist_mapping))

        return nist_cis_mapping   

    def map_controls_to_techniques(self):
        """
        Create a mapping between the prioritized ATT&CK techniques and the NIST controls.
        Goal is to calculate how many NIST controls associated with a given technique.
        Results are written to a file
        """
        attack_priorities_with_nist = []
        # Start with the ATT&CK techniques
        for priority in self.attack_priorities:
            # Isolate the technique
            rank = priority[0]
            technique = priority[1].upper()
            related_controls = set()
            # Search through the ATT&CK <-> NIST mappings for matches
            for mapping in self.attack_nist_mapping:
                control = mapping[0].upper()
                if control in self.new_cjis_nist_controls:
                    if technique in mapping:
                        # Add all of the controls related to a technique to the set()
                        related_controls.add(control)

            # Prepare Data    
            if related_controls:
                count = len(related_controls)

                # The details can be to much, so allow the user to decide if they care to see them.
                if self.include_details:
                    header = ("Priority (Assigned by MITRE/ATT&CK)", "Technique", "Number of Mapped Controls", "Controls")
                    related_controls = "|".join(related_controls)
                    attack_priorities_with_nist.append((rank, technique, count, related_controls))
                else:
                    header = ("Priority (Assigned by MITRE/ATT&CK)", "Technique", "Number of Mapped Controls (NIST 800-53)")
                    attack_priorities_with_nist.append((rank, technique, count))
                # Sort based on the third item in the tuples (count)
                attack_priorities_with_nist.sort(key=lambda x: x[0])

        # Write CSV file
        with open(self.attack_priorities_with_nist_file, 'w', newline='') as file:
            writer = csv.writer(file)
            writer.writerow(header)
            for row in attack_priorities_with_nist:
                writer.writerow(row)

    def map_techniques_to_controls(self):
        """
        Use the mapping between the NIST controls and ATT&CK.
        Goal is to calculate how many techniques are associated with each control. 
        Results are written to a file.
        """
        # Consolidate each NIST control into dictionaries,
        # Provides an easier view to how many techniques are associated with a control. 
        nist_with_mappings = []
        controls_and_techniques = {}
        for mapping in self.attack_nist_mapping:
            control = mapping[0].upper()
            # Only work on controls we care about
            if control in self.new_cjis_nist_controls:
                # Create a blank set() on the first instance of the control. 
                if control not in controls_and_techniques:
                    related_techiques = set()
                    related_cis_controls = set()
                    new_dict = {
                        f'{control}': {
                            'ATT&CK_techniques': related_techiques,
                            'cis_controls': related_cis_controls
                        }
                    }                  
                    controls_and_techniques.update(new_dict)

                # Isolate the technique
                technique = mapping[1].upper()
                
                # Create the mapping by adding the technique to our set()
                controls_and_techniques[f'{control}']['ATT&CK_techniques'].add(technique)
        
        for mapping in self.nist_cis_mapping:
            control = mapping[1]
            if control in self.new_cjis_nist_controls:
                # Isolate the CIS controls
                cis_control = mapping[0]

                # Create the mapping by adding the control to our set()
                controls_and_techniques[f'{control}']['cis_controls'].add(cis_control)


        # Prepare data
        for control, techniques in controls_and_techniques.items():
            technique_count = len(techniques['ATT&CK_techniques'])
            
            # Bunch of painful steps to get the CIS controls sorted and formated
            cis_controls_int = [int(x) for x in list(techniques['cis_controls'])] # Convert all items in the set() to integers
            cis_controls_int_sorted = sorted(cis_controls_int) # Sort the integers
            organized_cis_controls = []
            for i in cis_controls_int_sorted:
                noted_control = f'CIS-{i}' # Add CIS- to make it more apparent in the results
                organized_cis_controls.append(noted_control)        
            all_cis_controls = "|".join(organized_cis_controls) # covert to list, sorted, and join

            # The details can be to much, so allow the user to decide if they care to see them.
            if self.include_details:
                header = ("Control", "Number of ATT&CK Techniques & Sub-Techniques Mapped", "Related CIS Controls", "Techniques")
                all_techniques = "|".join(techniques['ATT&CK_techniques'])
                nist_with_mappings.append((control, technique_count, all_cis_controls, all_techniques))
            else:
                header = ("Control", "Number of ATT&CK Techniques & Sub-Techniques Mapped", "Related CIS Controls")
                nist_with_mappings.append((control, technique_count, all_cis_controls))
            
            # Sorted based on the second item in the tuples (technique_count)
            nist_with_mappings.sort(key=lambda x: x[1], reverse=True)

        # Write CSV file
        with open(self.nist_with_techniques_file, 'w', newline='') as file:
            writer = csv.writer(file)
            writer.writerow(header)
            for row in nist_with_mappings:
                writer.writerow(row)

def main():
    pc = PrioritizeControls("config.yaml")
    pc.map_controls_to_techniques()
    pc.map_techniques_to_controls()

if __name__ == main():
    main()
